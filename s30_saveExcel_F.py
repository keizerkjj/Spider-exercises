import openpyxl
from selenium import webdriver
from lxml import etree
from openpyxl import Workbook
'''
wb = Workbook()
sheet = wb.get_active_sheet()
ws = wb.active
'''



def save_to_Excel(name):
    wb = openpyxl.Workbook()
    ws =wb.active
    ws.cell(row=1, column=1).value = '书名'
    '''
    ws.cell(row=1, column=2).value = '评分'
    '''
    for i in range(25):
        result = [name[i]]
        ws.append(result)
    wb.save("douban.xlsx")

def get_web(url):
    driver = webdriver.Chrome()
    driver.get(url)

    print('waitting for .......')
    #time.sleep(10)
    print('waitting done .......')
    driver.set_page_load_timeout(5)

    driver.save_screenshot('douban_reader.png')


    fn = 'douban_reader.html'
    with open(fn, 'w', encoding='utf-8') as f:
        f.write(driver.page_source)

    content_parse(fn)
    driver.quit()

def content_parse(fn):

    name = []

    with open(fn, 'r', encoding='utf-8') as f:
        html = f.read()


    # 生成xml树
    tree = etree.HTML(html)

    #查找book
    books = tree.xpath('//div[@class="item-root"]')
    num0 = 2
    book_name = tree.xpath(".//div[@class='title']/a")
    for i in range(len(book_name)):
        name.append(book_name[i])

    save_to_Excel(name)

    for book in books:
        book_name = book.xpath(".//div[@class='title']/a")
        book_hrefs = book.xpath(".//div[@class='title']/a/@href")


        book_list.append(book_name)


        book_name = book_name[0].text

        '''
        bookname = []
        for i in book_name:
            bookname.append(i)
        for i in bookname:
            ws.cell(row=num0, column=1).value = i

            num0 = num0 + 1
            print(num0)
        '''

        print(book_name)
        print(book_hrefs[0])



        for book_href in book_hrefs:
            driver = webdriver.Chrome()
            driver.get(book_href)
            print('2nd web waitting for .......')
            driver.set_page_load_timeout(5)
            #time.sleep(10)
            print('2nd web waitting done .......')
            driver.save_screenshot('book_href.png')

            book = 'books.html'
            with open(book, 'w', encoding='utf-8') as b:
                b.write(driver.page_source)

            parse_detail(book)

            driver.quit()



def parse_detail(book):
    rating = []

    with open(book, 'r', encoding='utf-8') as b:
        bookhtml = b.read()
    tree = etree.HTML(bookhtml)

    rating_1 = tree.xpath('//div[@class="rating_wrap clearbox"]/div[@class="rating_self clearfix"]/strong')
    '''
    for i in range(len(rating_1)):
        rating.append(rating_1[i])
    '''


    if (len(rating_1)):
            '''
            rat = []
            
            for i in rating:
                rat.append(i)
            for i in rat:
                ws.cell(row=num1, column=2).value = i

                num1 = num1 + 1
                print(num1) 
            '''
            rating_1 = rating_1[0].text

            print(rating_1)





if __name__ == '__main__':
    urls = ['https://book.douban.com/subject_search?search_text=python&cat=1001&start={}'.format(str(i)) for i in range(0, 15, 15)]
    for url in urls:
        print(url)
        get_web(url)
